#!/usr/bin/env python3
"""
warranty_scan.py — Quét đơn còn bảo hành từ market_orders.json

Modes:
  --mode preview       — tính toán + trả JSON, KHÔNG ghi Sheet
  --mode create-sheet  — tính toán + tạo tab Google Sheet mới + lưu lịch sử
  --mode export-xlsx   — xuất file XLSX từ lịch sử (--scan-id required)
  --mode export-csv    — xuất file CSV từ lịch sử (--scan-id required)
"""
import argparse, json, os, re, sys, uuid
from datetime import datetime, date, timedelta
from typing import Optional

DATA_DIR = os.environ.get("DATA_DIR", os.path.join(os.path.dirname(os.path.abspath(__file__)), "data"))

# ─────────────────────────────── JSON helpers ────────────────────────────────

def _load_json(name, fallback=None):
    p = os.path.join(DATA_DIR, f"{name}.json")
    if not os.path.exists(p):
        return fallback
    try:
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return fallback

def _save_json(name, data):
    os.makedirs(DATA_DIR, exist_ok=True)
    p = os.path.join(DATA_DIR, f"{name}.json")
    with open(p, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ─────────────────────────────── Amount parsing ──────────────────────────────

def _parse_amount(s) -> int:
    """'350.000đ' | '350,000' | '350000' → 350000"""
    if not s:
        return 0
    s = str(s).replace("đ", "").replace(" ", "").replace(",", "")
    # Nếu có dấu '.' — kiểm tra xem dấu thập phân hay phân cách hàng nghìn
    # Giá VND thường không có phần thập phân, nên '.' là phân cách nghìn
    s = s.replace(".", "")
    try:
        return int(float(s))
    except ValueError:
        return 0

def _fmt_amount(n: int) -> str:
    """350000 → '350.000'"""
    return f"{n:,}".replace(",", ".")

# ─────────────────────────────── Date parsing ────────────────────────────────

def _parse_date_str(s) -> Optional[date]:
    """
    Hỗ trợ nhiều định dạng:
      - "HH:MM:SS dd/MM/yyyy"
      - "dd/MM/yyyy"
      - ISO "yyyy-MM-ddT..." (UTC offset OK)
    """
    if not s:
        return None
    s = str(s).strip()
    # ISO
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except ValueError:
        pass
    # "HH:MM:SS dd/MM/yyyy"
    m = re.match(r"\d{1,2}:\d{2}:\d{2}\s+(\d{2}/\d{2}/\d{4})", s)
    if m:
        try:
            return datetime.strptime(m.group(1), "%d/%m/%Y").date()
        except ValueError:
            pass
    # "dd/MM/yyyy"
    m = re.match(r"(\d{2}/\d{2}/\d{4})", s)
    if m:
        try:
            return datetime.strptime(m.group(1), "%d/%m/%Y").date()
        except ValueError:
            pass
    return None

def _fmt_date(d) -> str:
    """date | str → 'dd/MM/yyyy'"""
    if isinstance(d, str):
        parsed = _parse_date_str(d)
        if parsed:
            return parsed.strftime("%d/%m/%Y")
        return d
    if hasattr(d, "strftime"):
        return d.strftime("%d/%m/%Y")
    return str(d)

# ─────────────────────────────── Account parsing ─────────────────────────────

def _parse_account_fields(content: str) -> dict:
    """
    Tách tài khoản đã giao:
      'email@x.com / password / 2fa | Giao: ...'  → email/password/2fa
      'email@x.com | password | 2fa | Giao: ...'  → email/password/2fa
      Anything else                                → raw only
    """
    if not content:
        return {"email": "", "password": "", "twofa": "", "raw": ""}
    raw = content
    # Bỏ phần " | Giao: ..." ở cuối
    account_part = re.split(r"\s*\|\s*Giao\s*:", content, maxsplit=1)[0].strip()

    parts: list[str] = []
    if " / " in account_part:
        parts = [p.strip() for p in account_part.split(" / ")]
    elif "|" in account_part:
        parts = [p.strip() for p in account_part.split("|")]

    email    = parts[0] if len(parts) > 0 else ""
    password = parts[1] if len(parts) > 1 else ""
    twofa    = parts[2] if len(parts) > 2 else ""

    # Nếu "email" không có @ thì không phải định dạng email|pass|2fa
    if email and "@" not in email:
        email = password = twofa = ""

    return {"email": email, "password": password, "twofa": twofa, "raw": raw}

# ─────────────────────────────── Presets ─────────────────────────────────────

PRESETS: dict = {
    "chatgpt_30d": {
        "label":           "ChatGPT Plus BHF 30D",
        "warranty_days":   30,
        "include":         ["chatgpt", "chat gpt", "gpt plus", "chatgpt plus"],
        "warranty_include": ["30d", "30 d", "30 ngày", "bh30d", "bh 30d",
                             "bảo hành 30 ngày", "1 tháng bhf"],
        "exclude":         ["api", "codex", "token", "credit", "k12", "edu",
                            "bh 2", "bh2", "2 ngày", "bh 24h", "bh24h", "24h",
                            "1 ngày", "kbh", "không bảo hành", "no warranty"],
        "sheet_prefix":    "QUET_CHATGPT_BH30D",
    },
    "grok_super": {
        "label":           "Grok Super BHF",
        "warranty_days":   30,
        "include":         ["grok super", "grok"],
        "warranty_include": [],
        "exclude":         [],
        "sheet_prefix":    "QUET_GROK_SUPER",
    },
}

def _matches_preset(product_name: str, preset_key: str) -> tuple:
    """(True, '') | (False, reason)"""
    if preset_key not in PRESETS:
        return False, f"Preset không tồn tại: {preset_key}"
    p = PRESETS[preset_key]
    name_lc = product_name.lower()

    for kw in p["exclude"]:
        if kw in name_lc:
            return False, f"Loại trừ: chứa '{kw}'"

    if not any(kw in name_lc for kw in p["include"]):
        return False, "Không khớp tên sản phẩm"

    if p["warranty_include"] and not any(kw in name_lc for kw in p["warranty_include"]):
        return False, "Không khớp thời hạn bảo hành"

    return True, ""

VALID_STATUSES = {"completed", "success", "done", "hoàn tất", "hoàn thành"}

def _valid_status(status: str) -> bool:
    return (status or "").lower().strip() in VALID_STATUSES

# ─────────────────────────────── Core scan ───────────────────────────────────

def _get_start_date(order: dict) -> tuple:
    """(date | None, field_name)"""
    for field in ["completed_at", "payment_at", "delivered_at", "created_at_raw"]:
        d = _parse_date_str(order.get(field, ""))
        if d:
            return d, field
    return None, ""

def scan_orders(preset_key: str, scan_date: date, warranty_days: int,
                refund_mode: str, refund_price_fixed: int = 0) -> dict:
    """
    Trả về:
      stats, qualified (còn BH + đã dedup), expired, excluded, errors, duplicates
    """
    orders_map = _load_json("market_orders", {}) or {}
    orders = list(orders_map.values())

    qualified_raw: list = []
    expired:       list = []
    excluded:      list = []
    errors:        list = []

    for order in orders:
        oid     = order.get("order_id", "")
        product = order.get("product_name", "")

        matches, reason = _matches_preset(product, preset_key)
        if not matches:
            excluded.append({"order_id": oid, "product": product, "reason": reason})
            continue

        if not _valid_status(order.get("status", "")):
            excluded.append({"order_id": oid, "product": product,
                              "reason": f"Trạng thái: {order.get('status', '?')}"})
            continue

        start_date, date_field = _get_start_date(order)
        if not start_date:
            errors.append({"order_id": oid, "product": product, "reason": "Thiếu ngày mua/hoàn tất"})
            continue

        if start_date > scan_date:
            errors.append({"order_id": oid, "product": product,
                           "reason": f"Ngày mua ({start_date}) > ngày quét ({scan_date})"})
            continue

        expire_date = start_date + timedelta(days=warranty_days)
        days_used   = (scan_date - start_date).days
        days_left   = warranty_days - days_used

        sell_price   = _parse_amount(order.get("sell_price", 0))
        source_price = _parse_amount(order.get("price", 0))

        if refund_mode == "fixed":
            refund_amount = refund_price_fixed
        else:
            refund_amount = round(sell_price / warranty_days * days_left) if sell_price > 0 and warranty_days > 0 else 0

        acct = _parse_account_fields(order.get("content", ""))

        entry = {
            "order_id":    oid,
            "seller":      order.get("seller", ""),
            "buyer":       order.get("buyer", ""),
            "product_name": product,
            "account_raw": acct["raw"],
            "email":       acct["email"],
            "password":    acct["password"],
            "twofa":       acct["twofa"],
            "start_date":  start_date.isoformat(),
            "expire_date": expire_date.isoformat(),
            "warranty_days": warranty_days,
            "days_used":   days_used,
            "days_left":   days_left,
            "sell_price":  sell_price,
            "source_price": source_price,
            "refund_amount": refund_amount,
            "refund_status": "Chờ hoàn",
            "note":        "",
        }

        if days_left > 0:
            qualified_raw.append(entry)
        else:
            expired.append(entry)

    # Deduplicate
    qualified, duplicates = _deduplicate(qualified_raw)

    # Sort: oldest → newest
    qualified.sort(key=lambda x: x["start_date"])

    total_refund = sum(e["refund_amount"] for e in qualified)

    return {
        "stats": {
            "total_scanned":   len(orders),
            "total_matched":   len(qualified) + len(duplicates) + len(expired),
            "total_qualified": len(qualified),
            "total_expired":   len(expired),
            "total_excluded":  len(excluded),
            "total_errors":    len(errors),
            "total_duplicates": len(duplicates),
            "total_refund":    total_refund,
        },
        "qualified":  qualified,
        "expired":    expired,
        "excluded":   excluded,
        "errors":     errors,
        "duplicates": duplicates,
    }

def _deduplicate(orders: list) -> tuple:
    """
    Ưu tiên email > account_raw > order_id.
    Giữ bản mới nhất (start_date lớn nhất).
    Returns (unique, duplicates)
    """
    # Mới nhất trước để giữ khi trùng
    srt = sorted(orders, key=lambda x: x["start_date"], reverse=True)
    seen_email   = {}
    seen_account = {}
    seen_order   = {}
    unique    = []
    dups      = []

    for e in srt:
        email   = (e.get("email") or "").strip().lower()
        acct    = (e.get("account_raw") or "").strip()
        oid     = e.get("order_id", "")
        is_dup  = False
        reason  = ""

        if email and email in seen_email:
            is_dup = True; reason = f"Trùng email: {email}"
        elif acct and acct in seen_account:
            is_dup = True; reason = "Trùng tài khoản đã giao"
        elif oid and oid in seen_order:
            is_dup = True; reason = f"Trùng mã đơn: {oid}"

        if is_dup:
            dups.append({**e, "dup_reason": reason})
        else:
            unique.append(e)
            if email:   seen_email[email]   = True
            if acct:    seen_account[acct]  = True
            if oid:     seen_order[oid]     = True

    return unique, dups

# ─────────────────────────────── Google Sheets ───────────────────────────────

def _get_sheets_client():
    import gspread
    sa_json_str = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if not sa_json_str:
        sa_file = os.path.join(DATA_DIR, "google_sa.json")
        if os.path.exists(sa_file):
            with open(sa_file, encoding="utf-8") as f:
                sa_json_str = f.read().strip()
    if not sa_json_str:
        raise RuntimeError("Không tìm thấy Google Service Account credentials")
    sa_dict = json.loads(sa_json_str)
    return gspread.service_account_from_dict(sa_dict)

def _build_tab_name(preset_key: str, scan_date: date, ss) -> str:
    """Tạo tên tab, thêm _HHMM nếu đã tồn tại."""
    prefix  = PRESETS[preset_key]["sheet_prefix"]
    base    = f"{prefix}_{scan_date.strftime('%Y-%m-%d')}"
    existing = {ws.title for ws in ss.worksheets()}
    if base not in existing:
        return base
    suffix = datetime.now().strftime("%H%M")
    return f"{base}_{suffix}"

def create_sheet(scan_result: dict, preset_key: str, scan_date: date) -> dict:
    """
    Tạo tab mới trong Google Sheets và ghi dữ liệu đợt quét.
    Returns { ok, sheet_name, spreadsheet_url, error? }
    """
    cfg = _load_json("sheets_config", {}) or {}
    spreadsheet_id = cfg.get("spreadsheet_id", "")
    if not spreadsheet_id:
        return {"ok": False, "error": "Chưa cấu hình spreadsheet_id trong sheets_config.json"}

    try:
        client = _get_sheets_client()
        ss     = client.open_by_key(spreadsheet_id)
    except Exception as e:
        return {"ok": False, "error": f"Không mở được Google Sheet: {e}"}

    qualified = scan_result["qualified"]
    stats     = scan_result["stats"]
    preset    = PRESETS[preset_key]

    tab_name = _build_tab_name(preset_key, scan_date, ss)

    try:
        ws = ss.add_worksheet(title=tab_name, rows=max(10, len(qualified) + 10), cols=19)
    except Exception as e:
        return {"ok": False, "error": f"Không tạo được tab: {e}"}

    # ── Row 1: Tiêu đề ──────────────────────────────────────────────────────
    title_row = [f"🔍 Đợt quét: {tab_name} — {preset['label']}"]

    # ── Row 2: Thông tin tổng hợp ───────────────────────────────────────────
    summary_row = [
        f"Ngày quét: {_fmt_date(scan_date)}",
        f"Thời hạn BH: {preset['warranty_days']} ngày",
        f"Tổng tài khoản: {stats['total_qualified']}",
        f"Tổng tiền hoàn: {_fmt_amount(stats['total_refund'])}đ",
        f"Quét lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
    ]

    # ── Row 3: Trống ────────────────────────────────────────────────────────
    # ── Row 4: Headers ──────────────────────────────────────────────────────
    headers = [
        "STT", "Mã đơn", "Người bán", "Khách hàng", "Sản phẩm gốc",
        "Tài khoản đã giao", "Email", "Mật khẩu", "2FA",
        "Ngày mua", "Ngày hết hạn", "Tổng ngày BH", "Đã sử dụng", "Còn lại",
        "Giá bán", "Giá nguồn", "Tiền hoàn dự kiến", "Trạng thái hoàn", "Ghi chú",
    ]

    # ── Row 5+: Data ─────────────────────────────────────────────────────────
    data_rows = []
    for i, e in enumerate(qualified, 1):
        data_rows.append([
            i,
            e["order_id"],
            e["seller"],
            e["buyer"],
            e["product_name"],
            e["account_raw"],
            e["email"],
            e["password"],
            e["twofa"],
            _fmt_date(e["start_date"]),
            _fmt_date(e["expire_date"]),
            e["warranty_days"],
            e["days_used"],
            e["days_left"],
            e["sell_price"],
            e["source_price"],
            e["refund_amount"],
            e["refund_status"],
            e["note"],
        ])

    all_rows = [title_row, summary_row, [], headers] + data_rows

    try:
        ws.update(all_rows, "A1", value_input_option="USER_ENTERED")
    except Exception as e:
        return {"ok": False, "error": f"Không ghi dữ liệu được: {e}"}

    # Format: bold title + header, freeze row 4
    try:
        ws_id = ws.id
        ss.batch_update({"requests": [
            # Bold row 1 (title)
            {"repeatCell": {
                "range": {"sheetId": ws_id, "startRowIndex": 0, "endRowIndex": 1},
                "cell": {"userEnteredFormat": {"textFormat": {"bold": True, "fontSize": 12}}},
                "fields": "userEnteredFormat.textFormat",
            }},
            # Bold row 4 (header)
            {"repeatCell": {
                "range": {"sheetId": ws_id, "startRowIndex": 3, "endRowIndex": 4},
                "cell": {"userEnteredFormat": {
                    "textFormat": {"bold": True},
                    "backgroundColor": {"red": 0.267, "green": 0.408, "blue": 0.671},
                }},
                "fields": "userEnteredFormat(textFormat,backgroundColor)",
            }},
            # Header text color white
            {"repeatCell": {
                "range": {"sheetId": ws_id, "startRowIndex": 3, "endRowIndex": 4},
                "cell": {"userEnteredFormat": {"textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1}}}},
                "fields": "userEnteredFormat.textFormat.foregroundColor",
            }},
            # Freeze row 4
            {"updateSheetProperties": {
                "properties": {"sheetId": ws_id, "gridProperties": {"frozenRowCount": 4}},
                "fields": "gridProperties.frozenRowCount",
            }},
            # Auto-resize all columns
            {"autoResizeDimensions": {
                "dimensions": {"sheetId": ws_id, "dimension": "COLUMNS",
                               "startIndex": 0, "endIndex": 19},
            }},
            # Add filter on header row
            {"setBasicFilter": {
                "filter": {
                    "range": {
                        "sheetId": ws_id,
                        "startRowIndex": 3, "endRowIndex": 3 + len(data_rows) + 1,
                        "startColumnIndex": 0, "endColumnIndex": 19,
                    }
                }
            }},
        ]})
    except Exception as fmt_err:
        # Format failed — không critical
        pass

    spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit#gid={ws.id}"
    return {"ok": True, "sheet_name": tab_name, "spreadsheet_url": spreadsheet_url}

# ─────────────────────────────── Export helpers ──────────────────────────────

def _orders_to_rows(qualified: list) -> list:
    rows = [["STT","Mã đơn","Người bán","Khách hàng","Sản phẩm gốc",
             "Tài khoản đã giao","Email","Mật khẩu","2FA",
             "Ngày mua","Ngày hết hạn","Tổng ngày BH","Đã sử dụng","Còn lại",
             "Giá bán","Giá nguồn","Tiền hoàn dự kiến","Trạng thái hoàn","Ghi chú"]]
    for i, e in enumerate(qualified, 1):
        rows.append([
            i, e["order_id"], e["seller"], e["buyer"], e["product_name"],
            e["account_raw"], e["email"], e["password"], e["twofa"],
            _fmt_date(e["start_date"]), _fmt_date(e["expire_date"]),
            e["warranty_days"], e["days_used"], e["days_left"],
            e["sell_price"], e["source_price"], e["refund_amount"],
            e["refund_status"], e["note"],
        ])
    return rows

def export_xlsx(scan_id: str) -> str:
    """Tạo file XLSX từ lịch sử, trả về đường dẫn file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    history = _load_json("warranty_scan_history", []) or []
    entry   = next((h for h in history if h.get("scan_id") == scan_id), None)
    if not entry:
        raise ValueError(f"Không tìm thấy đợt quét: {scan_id}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = entry.get("sheet_name", "Quét BH")[:31]  # max 31 chars

    # Title row
    ws.append([f"🔍 Đợt quét: {entry.get('sheet_name', '')} — {entry.get('preset_label', '')}"])
    ws.append([
        f"Ngày quét: {entry.get('scan_date', '')}",
        f"Thời hạn BH: {entry.get('warranty_days', '')} ngày",
        f"Tổng tài khoản: {entry.get('total_qualified', '')}",
        f"Tổng tiền hoàn: {_fmt_amount(entry.get('total_refund', 0))}đ",
    ])
    ws.append([])

    rows = _orders_to_rows(entry.get("orders_snapshot", []))
    header_fill  = PatternFill("solid", fgColor="436AA8")
    header_font  = Font(bold=True, color="FFFFFF")
    for row_data in rows:
        ws.append(row_data)

    # Style header
    header_row = ws[4]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Bold title
    ws["A1"].font = Font(bold=True, size=12)
    ws.freeze_panes = "A5"

    # Auto-width (approximate)
    for col_idx in range(1, 20):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for row_obj in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row_obj:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 40))
        ws.column_dimensions[col_letter].width = max_len + 2

    out_dir = os.path.join(DATA_DIR, "warranty_exports")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"{scan_id}.xlsx")
    wb.save(out_path)
    return out_path

def export_csv(scan_id: str) -> str:
    """Tạo file CSV từ lịch sử, trả về đường dẫn file."""
    import csv

    history = _load_json("warranty_scan_history", []) or []
    entry   = next((h for h in history if h.get("scan_id") == scan_id), None)
    if not entry:
        raise ValueError(f"Không tìm thấy đợt quét: {scan_id}")

    out_dir = os.path.join(DATA_DIR, "warranty_exports")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"{scan_id}.csv")

    rows = _orders_to_rows(entry.get("orders_snapshot", []))
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    return out_path

# ─────────────────────────────── CLI main ────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Warranty Scan")
    parser.add_argument("--mode", required=True,
                        choices=["preview", "create-sheet", "export-xlsx", "export-csv"])
    parser.add_argument("--preset",       default="chatgpt_30d")
    parser.add_argument("--date",         default=None, help="yyyy-MM-dd, default=today")
    parser.add_argument("--warranty-days", type=int, default=None)
    parser.add_argument("--refund-mode",  default="sell_price",
                        choices=["sell_price", "fixed"])
    parser.add_argument("--refund-price", type=int, default=0)
    parser.add_argument("--scan-id",      default=None)
    args = parser.parse_args()

    # Export modes
    if args.mode == "export-xlsx":
        try:
            path = export_xlsx(args.scan_id)
            print(json.dumps({"ok": True, "path": path}))
        except Exception as e:
            print(json.dumps({"ok": False, "error": str(e)}))
        return

    if args.mode == "export-csv":
        try:
            path = export_csv(args.scan_id)
            print(json.dumps({"ok": True, "path": path}))
        except Exception as e:
            print(json.dumps({"ok": False, "error": str(e)}))
        return

    # Validate preset
    if args.preset not in PRESETS:
        print(json.dumps({"ok": False, "error": f"Preset không hợp lệ: {args.preset}"}))
        return

    preset_cfg = PRESETS[args.preset]

    # Scan date
    scan_date: date
    if args.date:
        try:
            scan_date = date.fromisoformat(args.date)
        except ValueError:
            print(json.dumps({"ok": False, "error": f"Ngày không hợp lệ: {args.date}"}))
            return
    else:
        scan_date = date.today()

    warranty_days = args.warranty_days if args.warranty_days else preset_cfg["warranty_days"]

    result = scan_orders(
        preset_key=args.preset,
        scan_date=scan_date,
        warranty_days=warranty_days,
        refund_mode=args.refund_mode,
        refund_price_fixed=args.refund_price,
    )

    if args.mode == "preview":
        print(json.dumps({"ok": True, **result}, ensure_ascii=False))
        return

    # create-sheet
    sheet_result = create_sheet(result, args.preset, scan_date)
    if not sheet_result["ok"]:
        print(json.dumps({"ok": False, "error": sheet_result["error"], **result}, ensure_ascii=False))
        return

    # Lưu lịch sử
    scan_id = str(uuid.uuid4())[:8].upper()
    history = _load_json("warranty_scan_history", []) or []
    history.insert(0, {
        "scan_id":        scan_id,
        "preset":         args.preset,
        "preset_label":   preset_cfg["label"],
        "scan_date":      scan_date.isoformat(),
        "warranty_days":  warranty_days,
        "refund_mode":    args.refund_mode,
        "refund_price_fixed": args.refund_price,
        "total_scanned":  result["stats"]["total_scanned"],
        "total_matched":  result["stats"]["total_matched"],
        "total_qualified": result["stats"]["total_qualified"],
        "total_refund":   result["stats"]["total_refund"],
        "sheet_name":     sheet_result["sheet_name"],
        "spreadsheet_url": sheet_result["spreadsheet_url"],
        "created_at":     datetime.now().isoformat(),
        "orders_snapshot": result["qualified"],
    })
    _save_json("warranty_scan_history", history)

    print(json.dumps({
        "ok":             True,
        "scan_id":        scan_id,
        "sheet_name":     sheet_result["sheet_name"],
        "spreadsheet_url": sheet_result["spreadsheet_url"],
        **result,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
